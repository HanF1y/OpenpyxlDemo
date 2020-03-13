from openpyxl import load_workbook

goal1 = goal2 = mfegoal = fegoal = wgoal = fagoal = 0
female = master = False
test1 = load_workbook('D:\code\python\other\比赛结果.xlsx')
test2 = load_workbook('D:\code\python\other\丙组积分表.xlsx')
sheet1 = test1.active
a = sheet1.max_column
b = sheet1.max_row
i = j = 1
for i in range(1,a):
    for j in range(1,b):
        c = sheet1.cell(column = i,row = j)
        temp = c.value
        if temp == "主队胜":
            wgoal+=1
        elif temp == "主队负":
            fagoal+=1
if wgoal == 1:
    goal1 = 1
    goal2 = 3
elif wgoal == 3|4:
    goal1 = 3
    goal2 = 1
else:
    for i in range(1, a):
        for j in range(1, b):
            c = sheet1.cell(column=i, row=j)
            temp = c.value
            if temp == "第一台（主将）":
                for p in range(i,a):
                    c = sheet1.cell(row=j,column=p)
                    temp = c.value
                    if temp == "主队胜":
                        master = True
                        goal1 = 2
                        goal2 = 1
                    elif temp =="主队负":
                        goal1 = 1
                        goal2 = 2
for i in range(1, a):
    for j in range(1, b):
        c = sheet1.cell(column=i, row=j)
        temp = c.value
        if temp == "女" :
            if i < 8:
                for p in range(i,a):
                    c = sheet1.cell(row=j,column=p)
                    temp = c.value
                    if temp == "主队胜":
                        mfegoal +=1
                        wgoal += 1
            else:
                for p in range(i,a):
                    c = sheet1.cell(row=j,column=p)
                    temp = c.value
                    if temp == "主队负":
                        fegoal +=1
                        fagoal +=1
for i in range(1, a):
    for j in range(1, b):
        c = sheet1.cell(column=i, row=j)
        temp = c.value
        if temp == "主队名称" :
            tname1 = sheet1.cell(column= i+1,row= j).value
        elif temp == "客队名称":
            tname2 = sheet1.cell(column= i+1,row= j).value
sheet2 = test2.active
a = sheet2.max_column
b = sheet2.max_row
for i in range(1, a):
    for j in range(1, b):
        c = sheet2.cell(column=i, row=j)
        temp = c.value
        if temp == "队名":
            for q in range(j,a):
                c = sheet2.cell(column=i, row=q)
                temp = c.value
                if temp == tname1:
                  sheet2.cell(column = i+1,row = q).value += goal1
                  sheet2.cell(column=i + 2, row=q).value += wgoal
                  if master:
                    sheet2.cell(column=i + 3, row=q).value +=1
                  sheet2.cell(column=i + 4, row=q).value += mfegoal
                elif temp == tname2:
                    sheet2.cell(column=i + 1, row=q).value += goal2
                    sheet2.cell(column=i + 2, row=q).value += fagoal
                    if not master:
                        sheet2.cell(column=i + 3, row=q).value += 1
                    sheet2.cell(column=i + 4, row=q).value += fegoal
test2.save('D:\code\python\other\丙组积分表第一轮.xlsx')